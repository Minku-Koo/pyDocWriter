import xlwings as xw
import pandas as pd
from pyxlsb import open_workbook as open_xlsb
import openpyxl as xl
from openpyxl.styles import PatternFill
from datetime import datetime
from tkinter import *
from tkinter.colorchooser import askcolor
from tkinter import filedialog


class TblColor:
    def __init__(self):
        guiTitle = "성능Tbl Compare"
        windowSize = "500x400+200+200"

        self.sheet_name = '성능 제어 파라메터 Table'

        self.diff_history_sheet_name = '변경이력'
        self.diff_history_col_n = 9

        doitBtn = 'WORK'

        self.source_path = ''
        self.target_path = ''
        self.color_code = ''


        self.window = Tk()
        self.window.title(guiTitle)
        self.window.geometry(windowSize)
        self.window.resizable(False, False)

        self.s_label = Label(self.window, text = 'Source Excel Tbl')
        self.s_label.pack()

        s_button = Button(self.window, text = 'Before', command = self.load_source)
        s_button.pack(pady=10)

        self.t_label = Label(self.window, text = 'Target Excel Tbl')
        self.t_label.pack()

        t_button = Button(self.window, text = 'After', command = self.load_target)
        t_button.pack(pady=10)

        c_button = Button(
            self.window,
            text='Select Color',
            command=self.change_color)
        c_button.pack(pady=14)
        
        self.c_label = Label(self.window, text = 'Select Color Code')
        self.c_label.pack()

        button = Button(self.window, text = doitBtn, command = self.clickEvent)
        button.pack(pady=10)

        self.result_lb = Label(self.window, text = 'Status')
        self.result_lb.pack(pady=20)

        self.asktome = Label(self.window, text = '문의 @minku.koo').pack(pady=10)

        self.window.mainloop()
        return 

    def load_source(self):
        self.source_path = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("Excel file", "*.xlsx"),('all files','*.*')))
        if self.source_path:
            self.s_label.config(text = self.source_path)
            self.result_lb.config(text='Status')
        return 

    def load_target(self):
        self.target_path = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("Excel file", "*.xlsx"),('all files','*.*')))
        if self.target_path:
            self.t_label.config(text = self.target_path)
            self.result_lb.config(text='Status')
        return 

    def change_color(self):
        colors = askcolor(title="Tkinter Color Chooser")
        if not colors[1]:
            return 
        self.color_code = colors[1][1:]
        self.c_label.config(text=self.color_code, bg='#'+self.color_code)
        return 

    def _change_status(self, txt):
        self.result_lb.config(text = txt)
        self.window.update()
        return 

    def clickEvent(self):
        if not self.source_path or not self.target_path or not self.color_code:
            self._change_status('모든 값을 입력해주셔야 합니다..')
            return 
        self._change_status('엑셀 읽는중...')

        _, ss = self._open_excel_openpyxl(self.source_path, self.sheet_name)
        tb, ts = self._open_excel_openpyxl(self.target_path, self.sheet_name)
        diff_tuple = self._compare_df(pd.DataFrame(ss.values), pd.DataFrame(ts.values))

        bgfill = PatternFill(start_color=self.color_code, end_color=self.color_code, fill_type='solid')

        self._change_status('변경점 색칠중...')

        for x, y in diff_tuple:
            ts.cell(row=y+1, column=x+1).fill = bgfill

        hs = tb[self.diff_history_sheet_name]
        hs = self._color_date(bgfill, hs, self.diff_history_col_n)
        tb.save(self.target_path)
        
        self._change_status('작업 완료!!')
        return 

    def _open_excel_openpyxl(self, filepath, sheetname):
        # Load workbook
        book = xl.load_workbook(filepath, "rb")
        # Access to a worksheet named 'no_header'
        sheet = book[sheetname]
        return book, sheet

    def _compare_df(self, df1, df2):
        result = []
        for x, _ in enumerate(df1.iterrows()):
            for y, _ in enumerate(df1.iloc[x]):
                if df1.iloc[x][y] == df2.iloc[x][y]:
                    continue
                result.append((y, x))
        return result

    def _color_date(self, cfill, sheet, col):
        now = datetime.now() # current date and time
        date_time = now.strftime("%m/%d/%Y, %H:%M:%S")

        # J열 따라 내려오면서 빈칸 찾기
        row = 1
        while True:
            if not sheet.cell(row=row, column=col).value:
                break
            row += 1

        # 빈칸 fill color and input datetime
        sheet.cell(row=row, column=col).fill = cfill
        sheet.cell(row=row, column=col).value = date_time
        return sheet


if __name__ == "__main__":

    tbl = TblColor()
